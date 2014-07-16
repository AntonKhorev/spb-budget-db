#!/usr/bin/env python3

from openpyxl import load_workbook
import tableWriters

def makeInvestmentTableReaderFromXlsxFile(inputFilename):
	wb=load_workbook(inputFilename)
	ws=wb.get_active_sheet()
	maxRow=ws.get_highest_row()
	for nRow in range(maxRow):
		if ws.cell(row=nRow,column=0).value=='ВСЕГО:':
			break
	else:
		raise Exception('table start not found')
	while nRow<maxRow:
		name=ws.cell(row=nRow,column=0).value
		if name.startswith('ЗАКАЗЧИК:') or name.startswith('ОТРАСЛЬ:'):
			nRow+=1
		yield [name]+[ws.cell(row=nRow,column=c).value for c in range(1,7)]
		nRow+=1

inputDirectory='../1-sources.out'
outputDirectory='../2-tables.out'

documentNumber=3574
tableWriters.InvestmentTableWriter(
	makeInvestmentTableReaderFromXlsxFile(inputDirectory+'/fincom/2014.0.p/pr23-2014-16.xlsx'),
	[2014,2015,2016]
).write(outputDirectory+'/2014.'+str(documentNumber)+'.23.investment.set(2014,2015,2016).csv')

documentNumber=3850
tableWriters.InvestmentTableWriter(
	makeInvestmentTableReaderFromXlsxFile(inputDirectory+'/fincom/2014.0.z/pr24_bd2014-16.xlsx'),
	[2014,2015,2016]
).write(outputDirectory+'/2014.'+str(documentNumber)+'.24.investment.set(2014,2015,2016).csv')
