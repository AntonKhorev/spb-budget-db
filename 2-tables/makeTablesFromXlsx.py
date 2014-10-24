#!/usr/bin/env python3

inputDirectory='../1-sources.out'
inputFincomDirectory=inputDirectory+'/fincom'
outputDirectory='../2-tables.out'
outputContentDirectory=outputDirectory+'/content'

from openpyxl import load_workbook
import tableWriters

def makeTableReaderFromXlsxFile(inputFilename,nAmountCols):
	def makeAmount(amount): # TODO delete?
		if type(amount) is int:
			return str(amount)+'.0'
		else:
			return str(amount)
	wb=load_workbook(inputFilename)
	ws=wb.get_active_sheet()
	maxRow=ws.get_highest_row()
	for nRow in range(maxRow):
		if ws.cell(row=nRow,column=0).value==1:
			break
	else:
		raise Exception('table start not found')
	for nRow in range(nRow,maxRow):
		yield [ws.cell(row=nRow,column=c).value for c in range(5)]+[makeAmount(ws.cell(row=nRow,column=c).value) for c in range(5,5+nAmountCols)]

dw=tableWriters.DepartmentTableWriter
sw=tableWriters.SectionTableWriter
for writer,fiscalYears,inputFilename,outputFilename in (
	(dw,[2014     ],'2014.0.p/pr03-2014-16.xlsx','2014.3574.3.department.set(2014).csv'),
	(dw,[2015,2016],'2014.0.p/pr04-2014-16.xlsx','2014.3574.4.department.set(2015,2016).csv'),
	(sw,[2014     ],'2014.0.p/pr05-2014-16.xlsx','2014.3574.5.section.set(2014).csv'),
	(sw,[2015,2016],'2014.0.p/pr06-2014-16.xlsx','2014.3574.6.section.set(2015,2016).csv'),
	(dw,[2014     ],'2014.0.z/pr03_bd2014-16.xlsx','2014.3850.3.department.set(2014).csv'),
	(dw,[2015,2016],'2014.0.z/pr04_bd2014-16.xlsx','2014.3850.4.department.set(2015,2016).csv'),
	(sw,[2014     ],'2014.0.z/pr05_bd2014-16.xlsx','2014.3850.5.section.set(2014).csv'),
	(sw,[2015,2016],'2014.0.z/pr06_bd2014-16.xlsx','2014.3850.6.section.set(2015,2016).csv'),
	(dw,[2014     ],'2014.1.z/pr02_bd2014-16_1izm.xlsx','2014.4752.2.department.set(2014).csv'),
	(dw,[2015,2016],'2014.1.z/pr17_bd2014-16_1izm.xlsx','2014.4752.17.department.diffset(3850,2015,2016).csv'),
	(sw,[2014     ],'2014.1.z/pr03_bd2014-16_1izm.xlsx','2014.4752.3.section.set(2014).csv'),
	(sw,[2015,2016],'2014.1.z/pr18_bd2014-16_1izm.xlsx','2014.4752.18.section.diffset(3850,2015,2016).csv'),
	(dw,[2015     ],'2015.0.p/pr03_2015.xlsx','2015.5208.3.department.set(2015).csv'),
	(dw,[2016,2017],'2015.0.p/pr04_2016-2017.xlsx','2015.5208.4.department.set(2016,2017).csv'),
	(sw,[2015     ],'2015.0.p/pr05_2015.xlsx','2015.5208.5.section.set(2015).csv'),
	(sw,[2016,2017],'2015.0.p/pr06-2016-2017.xlsx','2015.5208.6.section.set(2016,2017).csv'),
):
	writer(
		makeTableReaderFromXlsxFile(inputFincomDirectory+'/'+inputFilename,len(fiscalYears)),
		fiscalYears
	).write(outputContentDirectory+'/'+outputFilename)
