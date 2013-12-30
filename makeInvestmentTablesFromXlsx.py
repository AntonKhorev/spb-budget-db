#!/usr/bin/env python3

import decimal
import re
import csv
from openpyxl import load_workbook

def makeInvestmentTableReaderFromXlsxFile(inputFilename):
	def reader():
		wb=load_workbook(inputFilename)
		ws=wb.get_active_sheet()
		maxRow=ws.get_highest_row()
		for nRow in range(maxRow):
			if ws.cell(row=nRow,column=0).value=='ВСЕГО:':
				break
		else:
			raise Exception('table start not found')
		while nRow<maxRow:
			name=ws.cell(row=nRow,column=0).value
			if name.startswith('ЗАКАЗЧИК:') or name.startswith('ОТРАСЛЬ:'):
				nRow+=1
			yield [name]+[ws.cell(row=nRow,column=c).value for c in range(1,7)]
			nRow+=1
	return reader

class InvestmentTableWriter:
	def __init__(self,tableReader,years):
		self.rows=[]
		self.levelCols=[
			['grandTotal'],
			['departmentName','departmentTotal'],
			['branchName','branchTotal'],
			['recipientName','recipientTotal'],
			['projectName','districtNames','projectFirstYear','projectLastYear','projectDurationTotal','amount'],
		]
		yearRangeRe=re.compile(r'(\d{4})\s*-\s*(\d{4})')
		rows=[{'year':year} for year in years]
		def fixCase(s):
			s=s.replace('САНКТ-ПЕТЕРБУРГСКОМУ ГОСУДАРСТВЕННОМУ УНИТАРНОМУ ПРЕДПРИЯТИЮ ','СПБ ГУП ')
			s=s.replace('ОТКРЫТОМУ АКЦИОНЕРНОМУ ОБЩЕСТВУ ','ОАО ')
			return s
		def setLevel(toLevel):
			for level,cols in enumerate(self.levelCols):
				if toLevel<level:
					for col in cols:
						for row in rows:
							row.pop(col,None)
		def makeDecimal(v):
			return decimal.Decimal(v).quantize(decimal.Decimal('0.0'))
		def setColValue(k,v):
			for row in rows:
				row[k]=v
		for name,districtNames,yearRange,projectDurationTotal,*amounts in tableReader():
			def setColAmounts(k):
				for row,v in zip(rows,amounts):
					row[k]=makeDecimal(v)
			prefix='ВСЕГО:'
			if name.startswith(prefix):
				setLevel(0)
				setColAmounts('grandTotal')
				continue
			prefix='ЗАКАЗЧИК: '
			if name.startswith(prefix):
				departmentName=name[len(prefix):]
				setLevel(1)
				setColValue('departmentName',departmentName)
				setColAmounts('departmentTotal')
				continue
			prefix='ОТРАСЛЬ: '
			if name.startswith(prefix):
				branchName=name[len(prefix):]
				setLevel(2)
				setColValue('branchName',branchName)
				setColAmounts('branchTotal')
				continue
			prefix='БЮДЖЕТНЫЕ ИНВЕСТИЦИИ '
			if name.startswith(prefix):
				recipientName=fixCase(name[len(prefix):])
				setLevel(3)
				setColValue('recipientName',recipientName)
				setColAmounts('recipientTotal')
				continue
			setLevel(4)
			setColValue('projectName',name)
			setColValue('districtNames',districtNames)
			m=yearRangeRe.match(str(yearRange))
			if m:
				projectFirstYear=int(m.group(1))
				projectLastYear=int(m.group(2))
			else:
				projectFirstYear=projectLastYear=None if yearRange is None else int(yearRange)
			setColValue('projectFirstYear',projectFirstYear)
			setColValue('projectLastYear',projectLastYear)
			if projectDurationTotal:
				setColValue('projectDurationTotal',makeDecimal(projectDurationTotal))
			setColAmounts('amount')
			for row in rows:
				self.rows.append(row.copy())
	def write(self,outputFilename):
		cols=['year']+[col for cols in self.levelCols for col in cols]
		with open(outputFilename,'w',newline='',encoding='utf8') as file:
			writer=csv.writer(file,quoting=csv.QUOTE_NONNUMERIC)
			writer.writerow(cols)
			for row in self.rows:
				writer.writerow([row.get(col) for col in cols])

documentNumber=3574
InvestmentTableWriter(
	makeInvestmentTableReaderFromXlsxFile('fincom/2014.0.p/pr23-2014-16.xlsx'),
	[2014,2015,2016]
).write('tables/2014.0.p.'+str(documentNumber)+'.investment.csv')

documentNumber=3850
InvestmentTableWriter(
	makeInvestmentTableReaderFromXlsxFile('fincom/2014.0.z/pr24_bd2014-16.xlsx'),
	[2014,2015,2016]
).write('tables/2014.0.z.'+str(documentNumber)+'.investment.csv')
