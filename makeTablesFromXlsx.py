#!/usr/bin/env python3

from openpyxl import load_workbook
import tableWriters

def makeTableReaderFromXlsxFile(inputFilename,nAmountCols):
	def makeAmount(amount): # TODO delete?
		if type(amount) is int:
			return str(amount)+'.0'
		else:
			return str(amount)
	wb=load_workbook(inputFilename)
	ws=wb.get_active_sheet()
	maxRow=ws.get_highest_row()
	for nRow in range(maxRow):
		if ws.cell(row=nRow,column=0).value==1:
			break
	else:
		raise Exception('table start not found')
	for nRow in range(nRow,maxRow):
		yield [ws.cell(row=nRow,column=c).value for c in range(5)]+[makeAmount(ws.cell(row=nRow,column=c).value) for c in range(5,5+nAmountCols)]

# project
documentNumber=3574
tableWriters.DepartmentTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.0.p/pr03-2014-16.xlsx',1),
	[2014]
).write('tables/2014.0.p.'+str(documentNumber)+'.3.department.set(2014).csv')
tableWriters.DepartmentTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.0.p/pr04-2014-16.xlsx',2),
	[2015,2016]
).write('tables/2014.0.p.'+str(documentNumber)+'.4.department.set(2015,2016).csv')
tableWriters.SectionTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.0.p/pr05-2014-16.xlsx',1),
	[2014]
).write('tables/2014.0.p.'+str(documentNumber)+'.5.section.set(2014).csv')
tableWriters.SectionTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.0.p/pr06-2014-16.xlsx',2),
	[2015,2016]
).write('tables/2014.0.p.'+str(documentNumber)+'.6.section.set(2015,2016).csv')

# law
documentNumber=3850
tableWriters.DepartmentTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.0.z/pr03_bd2014-16.xlsx',1),
	[2014]
).write('tables/2014.0.z.'+str(documentNumber)+'.3.department.set(2014).csv')
tableWriters.DepartmentTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.0.z/pr04_bd2014-16.xlsx',2),
	[2015,2016]
).write('tables/2014.0.z.'+str(documentNumber)+'.4.department.set(2015,2016).csv')
tableWriters.SectionTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.0.z/pr05_bd2014-16.xlsx',1),
	[2014]
).write('tables/2014.0.z.'+str(documentNumber)+'.5.section.set(2014).csv')
tableWriters.SectionTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.0.z/pr06_bd2014-16.xlsx',2),
	[2015,2016]
).write('tables/2014.0.z.'+str(documentNumber)+'.6.section.set(2015,2016).csv')

# correction project
# wasn't published in xlsx format

# correction law
documentNumber=4752
diffsetDocumentNumber=3850
tableWriters.DepartmentTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.1.z/pr02_bd2014-16_1izm.xlsx',1),
	[2014]
).write('tables/2014.1.z.'+str(documentNumber)+'.2.department.set(2014).csv')
tableWriters.DepartmentTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.1.z/pr17_bd2014-16_1izm.xlsx',2),
	[2015,2016]
).write('tables/2014.1.z.'+str(documentNumber)+'.17.department.diffset('+str(diffsetDocumentNumber)+',2015,2016).csv')
tableWriters.SectionTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.1.z/pr03_bd2014-16_1izm.xlsx',1),
	[2014]
).write('tables/2014.1.z.'+str(documentNumber)+'.3.section.set(2014).csv')
tableWriters.SectionTableWriter(
	makeTableReaderFromXlsxFile('fincom/2014.1.z/pr18_bd2014-16_1izm.xlsx',2),
	[2015,2016]
).write('tables/2014.1.z.'+str(documentNumber)+'.18.section.diffset('+str(diffsetDocumentNumber)+',2015,2016).csv')
