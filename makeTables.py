#!/usr/bin/env python3

import decimal
import re
import csv
from openpyxl import load_workbook

class TableWriter:
	def __init__(self,inputFilename,years):
		self.years=years

		wb=load_workbook(inputFilename)
		ws=wb.get_active_sheet()
		maxRow=ws.get_highest_row()

		for nRow in range(maxRow):
			if ws.cell(row=nRow,column=0).value==1:
				break
		else:
			raise Exception('table start not found')

		self.rows=[{}]*len(years) # reserve first rows for totals
		self.processTable(ws,nRow,maxRow)

	def makeDecimalAmount(self,amount):
		if type(amount) is int:
			amount=str(amount)+'.0'
		else:
			amount=str(amount)
		return decimal.Decimal(amount)

	def makePrependRow(self):
		nRow=0
		def addRow(row):
			nonlocal nRow
			self.rows[nRow]=row
			nRow+=1
		return addRow

	def processTable(self,ws,nRow,maxRow):
		raise Exception('implement this')

	def getCols(self):
		raise Exception('implement this')

	def write(self,outputFilename):
		writer=csv.writer(open(outputFilename,'w',newline='',encoding='utf8'),quoting=csv.QUOTE_NONNUMERIC)
		cols=self.getCols()
		writer.writerow(cols)
		for row in self.rows:
			writer.writerow([row.get(col) for col in cols])

class DepartmentTableWriter(TableWriter):
	def processTable(self,ws,nRow,maxRow):
		departmentNameRe=re.compile(r'(?P<departmentName>.*?)\s*\((?P<departmentCode>...)\)')
		for nRow in range(nRow,maxRow):
			number,name,sectionCode,categoryCode,typeCode=(ws.cell(row=nRow,column=c).value for c in range(5))
			name=' '.join(name.split())
			amounts=[self.makeDecimalAmount(ws.cell(row=nRow,column=c).value) for c in range(5,5+len(self.years))]
			sectionCode=sectionCode[:2]+sectionCode[-2:]
			typeCode=str(typeCode)
			amountCol=None
			row={}
			addRow=self.rows.append
			if not number:
				addRow=self.makePrependRow()
				amountCol='yAmount'
			elif not sectionCode:
				m=departmentNameRe.match(name)
				row['departmentName']=departmentName=m.group('departmentName')
				row['departmentCode']=departmentCode=m.group('departmentCode')
				amountCol='ydAmount'
			elif not typeCode:
				row['departmentName']=departmentName
				row['departmentCode']=departmentCode
				row['sectionCode']=sectionCode
				row['categoryName']=categoryName=name
				row['categoryCode']=categoryCode
				amountCol='ydssccAmount'
			else:
				row['departmentName']=departmentName
				row['departmentCode']=departmentCode
				row['sectionCode']=sectionCode
				row['categoryName']=categoryName
				row['categoryCode']=categoryCode
				row['typeName']=typeName=name
				row['typeCode']=typeCode
				amountCol='ydsscctAmount'
			for year,amount in zip(self.years,amounts):
				r=row.copy()
				r['year']=year
				r[amountCol]=amount
				addRow(r)
			if not number:
				break
		else:
			raise Exception('no total line found')

	def getCols(self):
		return [
			'year',
			'departmentName','departmentCode',
			'sectionCode','categoryName','categoryCode',
			'typeName','typeCode',
			'yAmount','ydAmount','ydssccAmount','ydsscctAmount',
		]

class SectionTableWriter(TableWriter):
	def processTable(self,ws,nRow,maxRow):
		for nRow in range(nRow,maxRow):
			number,name,sectionCode,categoryCode,typeCode=(ws.cell(row=nRow,column=c).value for c in range(5))
			name=' '.join(name.split())
			amounts=[self.makeDecimalAmount(ws.cell(row=nRow,column=c).value) for c in range(5,5+len(self.years))]
			# sectionCode=sectionCode[:2]+sectionCode[-2:]
			if sectionCode[:2]!='  ':
				superSectionCode=sectionCode[:2]+'00'
			if sectionCode[-2:]!='  ':
				sectionCode=superSectionCode[:2]+sectionCode[-2:]
			else:
				sectionCode=''
			#
			typeCode=str(typeCode)
			amountCol=None
			row={}
			addRow=self.rows.append
			if not number:
				addRow=self.makePrependRow()
				amountCol='yAmount'
			elif not sectionCode:
				row['superSectionName']=superSectionName=name
				row['superSectionCode']=superSectionCode
				amountCol='ysAmount'
			elif not categoryCode:
				row['superSectionName']=superSectionName
				row['superSectionCode']=superSectionCode
				row['sectionName']=sectionName=name
				row['sectionCode']=sectionCode
				amountCol='yssAmount'
			elif not typeCode:
				row['superSectionName']=superSectionName
				row['superSectionCode']=superSectionCode
				row['sectionName']=sectionName
				row['sectionCode']=sectionCode
				row['categoryName']=categoryName=name
				row['categoryCode']=categoryCode
				amountCol='yssccAmount'
			else:
				row['superSectionName']=superSectionName
				row['superSectionCode']=superSectionCode
				row['sectionName']=sectionName
				row['sectionCode']=sectionCode
				row['categoryName']=categoryName
				row['categoryCode']=categoryCode
				row['typeName']=typeName=name
				row['typeCode']=typeCode
				amountCol='ysscctAmount'
			for year,amount in zip(self.years,amounts):
				r=row.copy()
				r['year']=year
				r[amountCol]=amount
				addRow(r)
			if not number:
				break
		else:
			raise Exception('no total line found')

	def getCols(self):
		return [
			'year',
			'superSectionName','superSectionCode',
			'sectionName','sectionCode',
			'categoryName','categoryCode',
			'typeName','typeCode',
			'yAmount','ysAmount','yssAmount','yssccAmount','ysscctAmount',
		]

DepartmentTableWriter('fincom/pr03-2014-16.xlsx',[2014]).write('tables/pr03-2014-16.csv')
DepartmentTableWriter('fincom/pr04-2014-16.xlsx',[2015,2016]).write('tables/pr04-2014-16.csv')
SectionTableWriter('fincom/pr05-2014-16.xlsx',[2014]).write('tables/pr05-2014-16.csv')
SectionTableWriter('fincom/pr06-2014-16.xlsx',[2015,2016]).write('tables/pr06-2014-16.csv')
