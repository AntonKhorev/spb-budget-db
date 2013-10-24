#!/usr/bin/env python3

import decimal
import re
import csv
from openpyxl import load_workbook

wb=load_workbook('fincom/pr03-2014-16.xlsx')
ws=wb.get_active_sheet()
maxRow=ws.get_highest_row()

for r in range(maxRow):
	c=ws.cell(row=r,column=0)
	if c.value==1:
		break
else:
	raise Exception('table start not found')

rows=[{}] # reserve first row for total
departmentNameRe=re.compile(r'(?P<departmentName>.*?)\s*\((?P<departmentCode>...)\)')

for r in range(r,maxRow):
	number,name,sectionCode,categoryCode,typeCode,amount=(ws.cell(row=r,column=c).value for c in range(6))
	sectionCode=sectionCode[:2]+sectionCode[-2:]
	typeCode=str(typeCode)
	if type(amount) is int:
		amount=str(amount)+'.0'
	else:
		amount=str(amount)
	amount=decimal.Decimal(amount)
	row={'year':2014}
	if not number:
		row['yearAmount']=amount
		rows[0]=row
		break
	elif not sectionCode:
		m=departmentNameRe.match(name)
		row['departmentName']=departmentName=m.group('departmentName')
		row['departmentCode']=departmentCode=m.group('departmentCode')
		row['yearDepartmentAmount']=amount
	elif not typeCode:
		row['departmentName']=departmentName
		row['departmentCode']=departmentCode
		row['sectionCode']=sectionCode
		row['categoryName']=categoryName=name
		row['categoryCode']=categoryCode
		row['yearDepartmentCategoryAmount']=amount
	else:
		row['departmentName']=departmentName
		row['departmentCode']=departmentCode
		row['sectionCode']=sectionCode
		row['categoryName']=categoryName
		row['categoryCode']=categoryCode
		row['typeName']=typeName=name
		row['typeCode']=typeCode
		row['yearDepartmentCategoryTypeAmount']=amount
	rows.append(row)
else:
	raise Exception('no total line found')

writer=csv.writer(open('tables/pr03-2014-16.csv','w',newline='',encoding='utf8'),quoting=csv.QUOTE_NONNUMERIC)
cols=[
	'year',
	'departmentName','departmentCode',
	'sectionCode','categoryName','categoryCode',
	'typeName','typeCode',
	'yearAmount','yearDepartmentAmount','yearDepartmentCategoryAmount','yearDepartmentCategoryTypeAmount',
]
writer.writerow(cols)
for row in rows:
	writer.writerow([row.get(col) for col in cols])
